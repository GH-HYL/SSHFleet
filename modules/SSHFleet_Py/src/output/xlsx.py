# -*- coding: utf-8 -*-
# SSHFleet Excel输出模块

import os
import re
import sys
from typing import Dict, List, Any
from posixpath import join as posix_join

import src.common.constants as color

from src.common.error_handler import error_and_exit_handling_decorator
from src.common.text_utils import clean_for_excel
from src.common.format_utils import format_conn_status, get_action_name, get_mode
from src.common.loader import SSHFleetConfig
from src.log import tlog


@error_and_exit_handling_decorator(
    "format_output_to_xlsx", "格式化输出结果到Excel文件失败", isexit=False
)
def format_output_to_xlsx(
    final_results: List[Dict[str, Any]],
    log_dir: str,
    args: Any,
    config: SSHFleetConfig,
) -> None:
    """
    功能：
        从结构化数据生成output.xlsx（3列格式：IP地址、事件类型、内容详情）

    参数:
        final_results: 结构化结果列表
        log_dir: 日志目录路径
        args: 命令行参数
        config: 配置对象
    返回:
        None
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    if not final_results:
        tlog.warning("结果列表为空，跳过 output.xlsx 生成")
        return

    # 配置输出路径
    os.makedirs(log_dir, exist_ok=True)
    output_path = posix_join(log_dir, config.paths.files.output_xlsx)

    # 创建Excel工作簿
    try:
        wb = Workbook()
        if wb.active is None:
            raise RuntimeError("Failed to create workbook with active sheet")

        ws: Worksheet = wb.active
        ws.title = "执行日志"
    except Exception as e:
        print(f"创建工作簿失败: {e}", file=sys.stderr)
        return

    # 设置样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    separator_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

    # 写入表头
    headers = ["IP地址", "事件类型", "内容详情"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if cell:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # 从结构化数据生成行
    row_idx = 2
    action = get_action_name(get_mode(args))

    for result in final_results:
        ip = result.get("ip", "未知IP")
        connect_success = result.get("connect_success", False)
        connect_cost_time = result.get("connect_cost_time", 0)
        exec_cost_time = result.get("exec_cost_time", 0)
        exit_code = result.get("exit_code", -1)
        output = clean_for_excel(result.get("output", ""))
        result_category = result.get("result_category", "未知")

        # 行1: 连接状态
        ws.cell(row=row_idx, column=1, value=ip)
        ws.cell(row=row_idx, column=2, value=format_conn_status(connect_success, connect_cost_time))
        row_idx += 1

        # 行2: 执行/上传状态
        exec_success = exit_code == 0
        exec_status = "成功" if exec_success else "失败"
        ws.cell(row=row_idx, column=1, value=ip)
        ws.cell(row=row_idx, column=2, value=f"{action}: {exec_status} - {exec_cost_time:.3f}s")
        row_idx += 1

        # 行3+: 输出内容（按行拆分）
        if output:
            for line in output.strip().split("\n"):
                if not line.strip():
                    continue
                ws.cell(row=row_idx, column=1, value=ip)
                ws.cell(row=row_idx, column=2, value="标准输出和错误输出")
                cell_content = ws.cell(row=row_idx, column=3, value=line.strip())
                if cell_content:
                    cell_content.alignment = Alignment(vertical="top", wrap_text=True)
                row_idx += 1

        # 行: 分类
        ws.cell(row=row_idx, column=1, value=ip)
        ws.cell(row=row_idx, column=2, value=f"分类: {result_category}")
        row_idx += 1

        # 分隔行
        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            if cell:
                cell.fill = separator_fill
        row_idx += 1

    # 设置列宽
    column_widths = [15, 20, 60]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_idx)
        if hasattr(ws, "column_dimensions"):
            ws.column_dimensions[col_letter].width = width

    # 冻结首行
    if hasattr(ws, "freeze_panes"):
        ws.freeze_panes = "A2"

    # 添加筛选
    if hasattr(ws, "auto_filter"):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"

    # 保存文件
    try:
        wb.save(output_path)
        tlog.success("格式化结构化数据到 output.xlsx 成功")
    except Exception as e:
        print(
            f"{color.COLOR_RED}[ERROR]{color.COLOR_RESET}{color.COLOR_YELLOW} [function:format_output_to_xlsx]{color.COLOR_RESET}保存Excel文件失败: {e}",
            file=sys.stderr,
        )
    finally:
        wb.close()


@error_and_exit_handling_decorator(
    "format_dict_list_to_xlsx", "格式化字典列表到Excel文件失败", isexit=False
)
def format_dict_list_to_xlsx(
    final_results: List[Dict[str, Any]], log_dir: str, config: SSHFleetConfig
) -> None:
    """
    功能：
        格式化字典列表为xlsx文件

    参数：
        final_results: 字典列表
        log_dir: 日志目录路径
        config: 配置对象
    返回值：
        None
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if not final_results:
        print(
            f"{color.COLOR_YELLOW}结果字典列表为空，未生成{config.paths.files.results_xlsx}{color.COLOR_RESET}"
        )
        tlog.error(f"结果字典列表为空，未生成{config.paths.files.results_xlsx}")
        return

    # clean_for_excel 清理字典列表中的ANSI转义序列和非法XML字符
    final_results = [{k: clean_for_excel(v) for k, v in item.items()} for item in final_results]

    # 生成输出路径
    os.makedirs(log_dir, exist_ok=True)
    dict_path = posix_join(log_dir, config.paths.files.results_xlsx)

    wb = None
    try:
        # 创建新工作簿
        wb = Workbook()
        if wb.active is None:
            raise RuntimeError("创建工作簿失败")

        ws = wb.active
        ws.title = "Results"

        # 基础样式配置
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        if not final_results:
            raise ValueError("Empty input data")

        headers = list(final_results[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据并计算列宽
        max_lengths = {header: len(header) for header in headers}

        for row_idx, row_data in enumerate(final_results, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if cell:  # 确保cell创建成功
                    cell.border = border
                    str_value = str(value)
                    if len(str_value) > max_lengths[header]:
                        max_lengths[header] = len(str_value)

        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            if hasattr(ws, "column_dimensions"):
                adjusted_width = min(max_lengths[header] + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # 添加自动筛选（如果工作表有效）
        if hasattr(ws, "auto_filter"):
            last_col = get_column_letter(len(headers))
            last_row = len(final_results) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        wb.save(dict_path)
        tlog.success("格式化results字典列表到Excel文件成功")

    except PermissionError as e:
        print(
            f"{getattr(color, 'COLOR_RED', '')}[ERROR]{getattr(color, 'COLOR_RESET', '')} "
            f"无法写入文件，请检查权限: {dict_path}\n异常类型：\n{type(e)}\n异常信息：\n{e}",
            file=sys.stderr,
        )
    except Exception as e:
        print(
            f"{getattr(color, 'COLOR_RED', '')}[ERROR]{getattr(color, 'COLOR_RESET', '')} "
            f"输出xlsx文件时出错\n异常类型：\n{type(e)}\n异常信息：\n{e}",
            file=sys.stderr,
        )
    finally:
        if wb is not None:
            wb.close()
