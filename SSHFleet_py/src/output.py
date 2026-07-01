# -*- coding: utf-8 -*-
# SSHexec 打印结果文件
# 该文件负责格式化输出结果，包括打印到终端和导出到Excel文件

# 系统或第三方模块
import argparse
import os
import re
import sys
from typing import Dict, List, Any
from posixpath import join as posix_join

import src.color as color

# 自定义模块
from src import utils
from src.toml import SSHExecConfig
from src.utils import tlog


@utils.error_and_exit_handling_decorator(
    "format_output_to_xlsx", "格式化输出结果到Excel文件失败", isexit=True
)
def format_output_to_xlsx(log_dir: str, config: SSHExecConfig) -> None:
    """
    功能：
        格式化output.txt文件内容到Excel文件

    参数:
        log_dir: 日志目录路径
        config: 配置对象
    返回:
        None
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    # 如果output.txt不存在，直接返回
    output_file = posix_join(log_dir, config.paths.files.output)
    if not os.path.exists(output_file):
        tlog.error(f"输出文本文件不存在，路径：{output_file}")
        return

    # 读取日志内容
    try:
        with open(output_file, "r", encoding="utf-8") as f:
            original_log_content = f.read()
            # 清理日志内容中的非法XML字符
            log_content = utils.clean_for_excel(original_log_content)

    except Exception as e:
        print(
            f"{color.COLOR_RED}[ERROR]{color.COLOR_RESET}{color.COLOR_YELLOW} [function:format_output_to_xlsx]{color.COLOR_RESET}读取日志文件失败: {e}",
            file=sys.stderr,
        )
        return

    # 配置输出路径
    os.makedirs(log_dir, exist_ok=True)
    output_path = posix_join(log_dir, config.paths.files.output_xlsx)

    # 创建Excel工作簿
    try:
        wb = Workbook()
        if wb.active is None:
            raise RuntimeError("Failed to create workbook with active sheet")

        ws: Worksheet = wb.active
        ws.title = "执行日志"
    except Exception as e:
        print(f"创建工作簿失败: {e}", file=sys.stderr)
        return

    # 设置样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["IP地址", "事件类型", "内容详情"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if cell:  # 确保cell对象存在
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # 处理日志内容
    current_ip = ""
    row_idx = 2
    lines = log_content.split("\n")

    for line in lines:
        if not line:
            continue

        # 处理══════行，转换为空白行
        # if line.startswith('══════'):
        if line.startswith("========================="):
            row_idx += 1
            continue

        # 匹配IP行
        ip_match = re.search(r"【(\d+\.\d+\.\d+\.\d+)】", line)
        if ip_match:
            current_ip = ip_match.group(1)
            content_start = line.find("】") + 1
            content = line[content_start:].strip()

            # 写入主记录
            ws.cell(row=row_idx, column=1, value=current_ip)
            ws.cell(row=row_idx, column=2, value=content)
            row_idx += 1
        elif current_ip:
            # 作为标准输出和错误输出内容
            ws.cell(row=row_idx, column=1, value=current_ip)
            ws.cell(row=row_idx, column=2, value="标准输出和错误输出")
            cell_content = ws.cell(row=row_idx, column=3, value=line.strip())

            # 设置内容单元格自动换行
            if cell_content:
                cell_content.alignment = Alignment(vertical="top", wrap_text=True)
            row_idx += 1

    # 设置列宽
    column_widths = [15, 20, 60]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_idx)
        if hasattr(ws, "column_dimensions"):
            ws.column_dimensions[col_letter].width = width

    # 冻结首行
    if hasattr(ws, "freeze_panes"):
        ws.freeze_panes = "A2"

    # 添加筛选
    if hasattr(ws, "auto_filter"):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"

    # 保存文件
    try:
        wb.save(output_path)
        return
    except Exception as e:
        print(
            f"{color.COLOR_RED}[ERROR]{color.COLOR_RESET}{color.COLOR_YELLOW} [function:format_output_to_xlsx]{color.COLOR_RESET}保存Excel文件失败: {e}",
            file=sys.stderr,
        )
        return
    finally:
        wb.close()


@utils.error_and_exit_handling_decorator(
    "format_dict_list_to_xlsx", "格式化字典列表到Excel文件失败", isexit=False
)
def format_dict_list_to_xlsx(
    final_results: List[Dict[str, Any]], log_dir: str, config: SSHExecConfig
) -> None:
    """
    功能：
        格式化字典列表为xlsx文件

    参数：
        final_results: 字典列表
        log_dir: 日志目录路径
        config: 配置对象
    返回值：
        None
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if not final_results:
        print(
            f"{color.COLOR_YELLOW}结果字典列表为空，未生成{config.paths.files.results_xlsx}{color.COLOR_RESET}"
        )
        tlog.error(f"结果字典列表为空，未生成{config.paths.files.results_xlsx}")
        return

    # clean_for_excel 清理字典列表中的ANSI转义序列和非法XML字符
    final_results = [{k: utils.clean_for_excel(v) for k, v in item.items()} for item in final_results]

    # 生成输出路径
    os.makedirs(log_dir, exist_ok=True)
    dict_path = posix_join(log_dir, config.paths.files.results_xlsx)

    wb = None
    try:
        # 创建新工作簿
        wb = Workbook()
        if wb.active is None:
            raise RuntimeError("创建工作簿失败")

        ws = wb.active
        ws.title = "Results"

        # 基础样式配置
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        if not final_results:
            raise ValueError("Empty input data")

        headers = list(final_results[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据并计算列宽
        max_lengths = {header: len(header) for header in headers}

        for row_idx, row_data in enumerate(final_results, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if cell:  # 确保cell创建成功
                    cell.border = border
                    str_value = str(value)
                    if len(str_value) > max_lengths[header]:
                        max_lengths[header] = len(str_value)

        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            if hasattr(ws, "column_dimensions"):
                adjusted_width = min(max_lengths[header] + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # 添加自动筛选（如果工作表有效）
        if hasattr(ws, "auto_filter"):
            last_col = get_column_letter(len(headers))
            last_row = len(final_results) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        wb.save(dict_path)

    except PermissionError as e:
        print(
            f"{getattr(color, 'COLOR_RED', '')}[ERROR]{getattr(color, 'COLOR_RESET', '')} "
            f"无法写入文件，请检查权限: {dict_path}\n异常类型：\n{type(e)}\n异常信息：\n{e}",
            file=sys.stderr,
        )
    except Exception as e:
        print(
            f"{getattr(color, 'COLOR_RED', '')}[ERROR]{getattr(color, 'COLOR_RESET', '')} "
            f"输出xlsx文件时出错\n异常类型：\n{type(e)}\n异常信息：\n{e}",
            file=sys.stderr,
        )
    finally:
        if wb is not None:
            wb.close()


@utils.error_and_exit_handling_decorator(
    "format_statistic_results_to_terminal",
    "格式化统计结果信息输出到终端失败",
    isexit=True,
)
def format_statistic_results_to_terminal(results_statistic: Dict) -> None:
    """
    功能：
        格式化统计结果信息输出到终端

    参数：
        results_statistic: 结果统计信息字典

    返回值：
        None
    """

    # print(f'\n')
    print("═" * 60)
    print(f"  总耗时： {results_statistic['global_cost_time']}  秒")
    if results_statistic["verify"] == "通过":
        print(
            f"  {color.COLOR_CYAN}节点总数:{color.COLOR_RESET} {results_statistic['nodeinofs_total']}  {color.COLOR_CYAN}完成总数：{color.COLOR_RESET}{results_statistic['results_total']}"
        )
    else:
        print(
            f"  {color.COLOR_CYAN}节点总数:{color.COLOR_RESET} {results_statistic['nodeinofs_total']}  {color.COLOR_CYAN}完成总数：{color.COLOR_RESET}{results_statistic['results_total']}  {color.COLOR_CYAN}总数校验：{color.COLOR_RESET}{results_statistic['verify_color']}{results_statistic['verify']}{color.COLOR_RESET}"
        )

    if results_statistic["fail_counts"] > 0:
        print(
            f"  {color.COLOR_GREEN}成功:{color.COLOR_RESET} {results_statistic['success_counts']}   {color.COLOR_RED}失败:{color.COLOR_RESET} {results_statistic['fail_counts']}"
        )
    else:
        print(
            f"  {color.COLOR_GREEN}成功:{color.COLOR_RESET} {results_statistic['success_counts']}"
        )

    if results_statistic["sorted_fail_categories"]:
        print(
            f'  {color.COLOR_RED}失败分类统计{color.COLOR_RESET} >>>  {"  ".join(f"{color.COLOR_YELLOW}{k}：{color.COLOR_RESET}{v}" for k, v in results_statistic["sorted_fail_categories"])}'
        )
    print("═" * 60)

    return


@utils.error_and_exit_handling_decorator(
    "format_statistic_results_to_report",
    "格式化统计结果信息输出到报告文件失败",
    isexit=True,
)
def format_statistic_results_to_report(
    results_statistic: Dict,
    log_dir: str,
    args: argparse.Namespace,
    config: SSHExecConfig,
) -> None:
    """
    功能：
        格式化统计结果信息输出到报告文件

    参数：
        results_statistic: 结果统计信息字典
        log_dir: 日志目录路径
        args: 命令行参数
        config: 配置对象

    返回值：
        None
    """
    report_file = posix_join(log_dir, config.paths.files.report)

    # 格式化执行命令内容
    args_set = sys.argv[1:]
    args_content = " ".join(args_set) if args_set else ""
    command = f"python3 {sys.argv[0]} {args_content}"

    with open(report_file, "w", encoding="utf-8") as f:
        f.write(
            "=============================执行结果统计报告=============================\n"
        )
        f.write(f'执行开始时间： {results_statistic["global_start_time"]}\n')
        f.write(f'执行结束时间： {results_statistic["global_stop_time"]}\n')
        f.write(f'执行耗时： {results_statistic["global_cost_time"]}  秒\n')
        f.write(f"\n【执行命令】 \n  {command}\n")
        f.write("\n【执行参数】\n")
        if args.c:
            f.write("  执行模式： 命令模式\n")
            f.write(f"  执行命令： {args.c}\n")
        if args.s:
            f.write("  执行模式： 脚本模式\n")
            f.write(f"  脚本路径： {args.s}\n")
        # if args.e:
        #     f.write(f"  环境变量： {args.e}\n")
        if args.u:
            f.write("  执行模式： 上传模式\n")
            f.write(f"  本地路径： {args.u}\n")
            f.write(f"  远程路径： {args.p}\n")
        if args.d:
            f.write("  执行模式： 下载模式\n")
            f.write(f"  远程路径： {args.d}\n")
            f.write(f"  本地路径： {args.p}\n")

        f.write(f"  CSV文件路径： {args.f}\n")
        f.write(f'  节点数量： {results_statistic["nodeinofs_total"]}\n')
        if args.c or args.s:
            f.write(f"  并发数值： {args.n}\n")
        if args.T:
            f.write(f"  连接超时： {args.T}s\n")
        if args.t:
            if args.c or args.s:
                f.write(f"  执行超时： {args.t}s\n")
            if args.u or args.d:
                f.write(f"  传输超时： {args.t}s\n")

        f.write("\n【结果统计】\n")
        f.write(f"  总耗时： {results_statistic['global_cost_time']}  秒\n")
        f.write(
            f"  节点总数: {results_statistic['nodeinofs_total']}  完成总数：{results_statistic['results_total']}  总数校验：{results_statistic['verify']}\n"
        )
        f.write(
            f"  成功: {results_statistic['success_counts']}    失败: {results_statistic['fail_counts']}\n"
        )

        if results_statistic["sorted_fail_categories"]:
            f.write(
                f'  失败分类统计 -→  {"  ".join(f"{k}：{v}" for k, v in results_statistic["sorted_fail_categories"])}\n'
            )

        f.write("\n【IP清单统计】\n")

        # 先输出失败分类（按IP数量升序排列）
        sorted_fail_items = sorted(
            results_statistic["category_ip_map"].items(), key=lambda x: len(x[1])
        )
        for category, ips in sorted_fail_items:
            f.write(f"\n{category}（{len(ips)}）：\n")
            for ip in ips:
                f.write(f"{ip}\n")

        # 最后输出成功分类
        if results_statistic["sorted_success_ips"]:
            f.write(
                f'\n{results_statistic["success_category"]}（{results_statistic["success_ips_count"]}）：\n'
            )
            for ip in results_statistic["sorted_success_ips"]:
                f.write(f"{ip}\n")
